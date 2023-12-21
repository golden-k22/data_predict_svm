# Code in file nn/two_layer_net_module.py
import torch
import pandas as pd
from openpyxl import load_workbook


class Feedforward(torch.nn.Module):
    def __init__(self, input_size, hidden_size, output_size):
        super(Feedforward, self).__init__()
        self.input_size = input_size
        self.hidden_size = hidden_size
        self.fc1 = torch.nn.Linear(self.input_size, self.hidden_size)
        self.relu = torch.nn.ReLU()
        self.fc2 = torch.nn.Linear(self.hidden_size, output_size)
        # self.sigmoid = torch.nn.Sigmoid()

    def forward(self, x):
        hidden = self.fc1(x)
        relu = self.relu(hidden)
        output = self.fc2(relu)
        # output = self.sigmoid(output)

        soft_max = torch.argmax(output, dim=1)
        soft_max = soft_max.type(torch.float32)
        return soft_max


# class TwoLayerNet(torch.nn.Module):
#     def __init__(self, D_in, H, D_out):
#         """
#         In the constructor we instantiate two nn.Linear modules and
#         assign them as
#         member variables.
#         """
#         super(TwoLayerNet, self).__init__()
#         self.linear1 = torch.nn.Linear(D_in, H)
#         # self.linear1 = torch.nn.LSTM(D_in, H)
#         self.linear2 = torch.nn.Linear(H, D_out)
#         # self.softmax = torch.nn.Softmax()
#
#     def forward(self, x):
#         """
#         In the forward function we accept a Tensor of input data and we must return
#         a Tensor of output data. We can use Modules defined in the constructor as
#         well as arbitrary (differentiable) operations on Tensors.
#         """
#         h_relu = self.linear1(x).clamp(min=0)
#         # h_relu = self.linear1(x)
#         y_pred = self.linear2(h_relu)
#         soft_max = torch.argmax(y_pred, dim=1)
#         soft_max = soft_max.type(torch.float32)
#
#         return soft_max

# N is batch size; D_in is input dimension;
# H is hidden dimension; D_out is output dimension.
N, D_in, H, D_out = 64, 17, 68, 100


def train():
    # Construct our model by instantiating the class defined above.
    model = Feedforward(D_in, H, D_out)

    # Construct our loss function and an Optimizer. The call to
    model.parameters()
    loss_fn = torch.nn.MSELoss(size_average=False)
    optimizer = torch.optim.SGD(model.parameters(), lr=1e-4)

    # get total data
    total_data = pd.DataFrame()

    xls = pd.ExcelFile('DataToShareMoreAccurate.xlsx')
    sheet_names = xls.sheet_names
    for i in range(1, len(sheet_names) - 1):
        # if i > 2:
        #     break
        df = pd.read_excel(xls, sheet_names[i])
        total_data = total_data.append(df)

    epoch = 500
    model.train()
    for i in range(epoch):
        total_data = total_data.sample(frac=1).reset_index(drop=True)
        for j in range(0, int(total_data.shape[0] / N) + 1):
            if j < int(total_data.shape[0] / N):
                train_df = total_data.iloc[j * N:(j + 1) * N, 3:20]
                finalScore = total_data.iloc[j * N:(j + 1) * N, :]['Final Score']
            else:
                train_df = total_data.iloc[total_data.shape[0] - N:total_data.shape[0], 3:20]
                finalScore = total_data.iloc[total_data.shape[0] - N:total_data.shape[0], :]['Final Score']


            x = torch.tensor(train_df.values)
            x = x.type(torch.float32)
            y = torch.tensor(finalScore.values)
            y = y.type(torch.float32)
            # Forward pass: Compute predicted y by passing x to the model
            y_pred = model(x)

            # Compute and print loss
            loss = loss_fn(y_pred, y)
            loss.requires_grad = True
            print("{}/{}".format(i, epoch), loss.item())

            # Zero gradients, perform a backward pass, and update the weights.
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
    torch.save(model, 'predict.pt')


def predict():
    # device = torch.device('cpu' if torch.cuda.is_available() else 'gpu')

    model = torch.load('predict.pt')
    # model.load_state_dict(torch.load('predict.pt'))

    # model = model.to(device) # Set model to gpu
    model.eval()

    xls = pd.ExcelFile('DataToShareMoreAccurate.xlsx')
    inputs = pd.read_excel(xls, 'ToPredict')
    # inputs = inputs.to(device) # You can move your input to gpu, torch defaults to cpu

    predicted_score = []
    # Run forward pass
    with torch.no_grad():

        for j in range(0, int(inputs.shape[0] / N) + 1):
            if j < int(inputs.shape[0] / N):
                df = inputs.iloc[j * N:(j + 1) * N, 3:20]
            else:
                df = inputs.iloc[inputs.shape[0] - N:inputs.shape[0], 3:20]

            x = torch.tensor(df.values)
            x = x.type(torch.float32)
            pred = model(x)

            # Do something with pred
            pred = pred.detach().cpu().numpy()
            end_index = N
            for row in pred:
                if j < int(inputs.shape[0] / N):
                    predicted_score.append(row)
                else:
                    unfilled_cnt = int(inputs.shape[0]) - len(predicted_score)
                    if unfilled_cnt == end_index:
                        predicted_score.append(row)
                    end_index -= 1

        wb = load_workbook('DataToShareMoreAccurate.xlsx')
        ws = wb['ToPredict']
        for i in range(len(predicted_score)):
            cell = 'U%d' % (i + 2)
            ws[cell] = predicted_score[i]

        wb.save('test.xlsx')


if __name__ == '__main__':
    train()
    # predict()
