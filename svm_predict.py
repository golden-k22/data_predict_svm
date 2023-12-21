import numpy as np
from openpyxl import load_workbook
from sklearn.svm import SVC, SVR
import pandas as pd
import pickle

EXCEL_FILE = 'DataToShareMoreAccurate.xlsx'
MODEL_NAME = 'model_all.pkl'


def train():
    # get total data
    total_data = pd.DataFrame()

    xls = pd.ExcelFile(EXCEL_FILE)
    sheet_names = xls.sheet_names
    for i in range(1, len(sheet_names) - 1):
        # if i > 5:
        #     break
        df = pd.read_excel(xls, sheet_names[i])
        is_nan = df.isnull().values.any()
        if is_nan == False:
            total_data = total_data.append(df)

    train_df = total_data.iloc[:, 3:20]
    finalScore = total_data.iloc[:, :]['Final Score']

    # svm_reg = SVR(gamma='auto', C=0.1, epsilon=0.2) # original model
    svm_reg = SVR(gamma='auto', C=100, epsilon=0.2)
    svm_reg.fit(train_df, finalScore)

    # clf = SVC(C=1, gamma=0.0001, kernel='linear')
    # clf.fit(train_df, finalScore)
    # save
    with open(MODEL_NAME, 'wb') as f:
        pickle.dump(svm_reg, f)


def predict():
    xls = pd.ExcelFile(EXCEL_FILE)
    inputs = pd.read_excel(xls, 'ToPredict')
    df = inputs.iloc[:, 3:20]
    # load
    with open(MODEL_NAME, 'rb') as f:
        clf2 = pickle.load(f)
    pred = clf2.predict(df)
    pred=np.rint(pred)
    print(pred)
    wb = load_workbook(EXCEL_FILE)
    ws = wb['ToPredict']
    for i in range(len(pred)):
        cell = 'U%d' % (i + 2)
        ws[cell] = pred[i]

    wb.save('test_all.xlsx')


# train()
predict()
